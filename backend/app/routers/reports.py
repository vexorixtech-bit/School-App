from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date, datetime
from app.database import get_db
from app.models.models import User, UserRole, Student, Teacher, Attendance, Fee, Exam, Result, Class
from app.auth.auth_handler import get_current_user
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from openpyxl import Workbook
from io import BytesIO
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def generate_pdf_response(buf, filename):
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

def generate_excel_response(buf, filename):
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

# --- Student Reports ---
@router.get("/students")
def get_student_report(
    class_id: Optional[int] = None, format: str = "pdf",
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    query = db.query(Student).filter(Student.is_active == True)
    if class_id:
        query = query.filter(Student.class_id == class_id)
    students = query.limit(1000).all()
    
    if format == "excel":
        data = [{
            "ID": s.student_id, "Name": f"{s.first_name} {s.last_name}",
            "DOB": s.date_of_birth, "Gender": s.gender,
            "Phone": s.phone, "Email": s.email
        } for s in students]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        
        buf = BytesIO()
        wb.save(buf)
        return generate_excel_response(buf, "student_report.xlsx")
    
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(letter))
    width, height = landscape(letter)
    
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width/2, height - 40, "Student Report")
    c.setFont("Helvetica", 10)
    
    y = height - 80
    c.drawString(30, y, "ID")
    c.drawString(120, y, "Name")
    c.drawString(280, y, "Class")
    c.drawString(360, y, "Gender")
    c.drawString(430, y, "Phone")
    c.drawString(540, y, "Email")
    c.line(30, y - 5, width - 30, y - 5)
    
    y -= 20
    for s in students:
        cls = db.query(Class).filter(Class.id == s.class_id).first()
        c.drawString(30, y, s.student_id)
        c.drawString(120, y, f"{s.first_name} {s.last_name}")
        c.drawString(280, y, cls.name if cls else "N/A")
        c.drawString(360, y, s.gender.value if hasattr(s.gender, 'value') else str(s.gender))
        c.drawString(430, y, s.phone or "N/A")
        c.drawString(540, y, s.email or "N/A")
        y -= 15
        if y < 40:
            c.showPage()
            y = height - 40
    
    c.save()
    return generate_pdf_response(buf, "student_report.pdf")

# --- Attendance Reports ---
@router.get("/attendance")
def get_attendance_report(
    class_id: Optional[int] = None, month: Optional[int] = None,
    year: Optional[int] = None, format: str = "pdf",
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    if not year: year = date.today().year
    if not month: month = date.today().month
    
    query = db.query(Attendance).filter(
        func.extract('year', Attendance.date) == year,
        func.extract('month', Attendance.date) == month
    )
    if class_id:
        query = query.filter(Attendance.class_id == class_id)
    
    records = query.limit(5000).all()
    total_days = len(set(r.date for r in records))
    
    # Per student summary
    students = db.query(Student).filter(Student.is_active == True)
    if class_id:
        students = students.filter(Student.class_id == class_id)
    
    data = []
    for student in students:
        student_records = [r for r in records if r.student_id == student.id]
        present = sum(1 for r in student_records if r.status == "present")
        absent = sum(1 for r in student_records if r.status == "absent")
        late = sum(1 for r in student_records if r.status == "late")
        data.append({
            "ID": student.student_id,
            "Name": f"{student.first_name} {student.last_name}",
            "Present": present,
            "Absent": absent,
            "Late": late,
            "Percentage": round((present / total_days * 100), 2) if total_days > 0 else 0
        })
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        buf = BytesIO()
        wb.save(buf)
        return generate_excel_response(buf, "attendance_report.xlsx")
    
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(letter))
    width, height = landscape(letter)
    
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width/2, height - 40, f"Attendance Report - {month}/{year}")
    c.setFont("Helvetica", 10)
    
    y = height - 80
    headers = ["ID", "Name", "Present", "Absent", "Late", "Percentage"]
    x_positions = [30, 120, 350, 420, 480, 550]
    for i, h in enumerate(headers):
        c.drawString(x_positions[i], y, h)
    c.line(30, y - 5, width - 30, y - 5)
    
    y -= 20
    for row in data:
        c.drawString(x_positions[0], y, row["ID"])
        c.drawString(x_positions[1], y, row["Name"][:25])
        c.drawString(x_positions[2], y, str(row["Present"]))
        c.drawString(x_positions[3], y, str(row["Absent"]))
        c.drawString(x_positions[4], y, str(row["Late"]))
        c.drawString(x_positions[5], y, f'{row["Percentage"]}%')
        y -= 15
    
    c.save()
    return generate_pdf_response(buf, "attendance_report.pdf")

# --- Fee Reports ---
@router.get("/fees")
def get_fee_report(
    class_id: Optional[int] = None, status: Optional[str] = None,
    from_date: Optional[str] = None, to_date: Optional[str] = None,
    format: str = "pdf", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    query = db.query(Fee)
    if class_id:
        query = query.join(Student).filter(Student.class_id == class_id)
    if status:
        query = query.filter(Fee.status == status)
    if from_date:
        query = query.filter(Fee.due_date >= date.fromisoformat(from_date))
    if to_date:
        query = query.filter(Fee.due_date <= date.fromisoformat(to_date))
    
    fees = query.order_by(Fee.due_date.desc()).limit(1000).all()
    
    data = [{
        "Receipt": f.receipt_no or "N/A",
        "Student": f"{db.query(Student).filter(Student.id == f.student_id).first().first_name if db.query(Student).filter(Student.id == f.student_id).first() else ''}",
        "Type": f.fee_type,
        "Amount": f.amount,
        "Paid": f.paid_amount,
        "Due Date": f.due_date,
        "Status": f.status.value if hasattr(f.status, 'value') else str(f.status)
    } for f in fees]
    
    total_collected = sum(d["Paid"] for d in data)
    total_pending = sum(d["Amount"] - d["Paid"] for d in data)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Fees"
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
            ws.append([])
            ws.append(["Total Collected", total_collected])
            ws.append(["Total Pending", total_pending])
        buf = BytesIO()
        wb.save(buf)
        return generate_excel_response(buf, "fee_report.xlsx")
    
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(letter))
    width, height = landscape(letter)
    
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width/2, height - 40, "Fee Report")
    c.setFont("Helvetica", 9)
    
    y = height - 80
    headers = ["Receipt", "Student", "Type", "Amount", "Paid", "Due Date", "Status"]
    x_pos = [30, 120, 250, 350, 420, 490, 560]
    for i, h in enumerate(headers):
        c.drawString(x_pos[i], y, h)
    c.line(30, y - 5, width - 30, y - 5)
    
    y -= 18
    for row in data:
        c.drawString(x_pos[0], y, str(row["Receipt"])[:12])
        c.drawString(x_pos[1], y, str(row["Student"])[:15])
        c.drawString(x_pos[2], y, str(row["Type"])[:12])
        c.drawString(x_pos[3], y, f'Rs.{row["Amount"]}')
        c.drawString(x_pos[4], y, f'Rs.{row["Paid"]}')
        c.drawString(x_pos[5], y, str(row["Due Date"]))
        c.drawString(x_pos[6], y, str(row["Status"]))
        y -= 15
    
    y -= 20
    c.setFont("Helvetica-Bold", 11)
    c.drawString(30, y, f"Total Collected: Rs.{total_collected}")
    c.drawString(300, y, f"Total Pending: Rs.{total_pending}")
    
    c.save()
    return generate_pdf_response(buf, "fee_report.pdf")

# --- Exam Reports ---
@router.get("/exams")
def get_exam_report(
    exam_id: Optional[int] = None, class_id: Optional[int] = None,
    format: str = "pdf", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    query = db.query(Exam)
    if class_id:
        query = query.filter(Exam.class_id == class_id)
    exams = query.all()
    
    data = []
    for exam in exams:
        results = db.query(Result).filter(Result.exam_id == exam.id)
        avg_marks = db.query(func.avg(Result.marks_obtained)).filter(Result.exam_id == exam.id).scalar() or 0
        total_students = results.count()
        passed = results.filter(Result.marks_obtained >= Result.max_marks * 0.35).count()
        
        data.append({
            "Exam": exam.name,
            "Type": exam.exam_type.value if hasattr(exam.exam_type, 'value') else str(exam.exam_type),
            "Avg Marks": round(float(avg_marks), 2),
            "Students": total_students,
            "Passed": passed,
            "Pass %": round((passed / total_students * 100), 2) if total_students > 0 else 0
        })
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Exams"
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        buf = BytesIO()
        wb.save(buf)
        return generate_excel_response(buf, "exam_report.xlsx")
    
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(letter))
    width, height = landscape(letter)
    
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width/2, height - 40, "Exam Report")
    c.setFont("Helvetica", 10)
    
    y = height - 80
    headers = ["Exam", "Type", "Avg Marks", "Students", "Passed", "Pass %"]
    x_pos = [30, 200, 320, 410, 490, 560]
    for i, h in enumerate(headers):
        c.drawString(x_pos[i], y, h)
    c.line(30, y - 5, width - 30, y - 5)
    
    y -= 20
    for row in data:
        c.drawString(x_pos[0], y, row["Exam"][:20])
        c.drawString(x_pos[1], y, row["Type"])
        c.drawString(x_pos[2], y, str(row["Avg Marks"]))
        c.drawString(x_pos[3], y, str(row["Students"]))
        c.drawString(x_pos[4], y, str(row["Passed"]))
        c.drawString(x_pos[5], y, f'{row["Pass %"]}%')
        y -= 15
    
    c.save()
    return generate_pdf_response(buf, "exam_report.pdf")
